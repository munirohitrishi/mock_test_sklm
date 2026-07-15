"""
Jyotirao Phule Study Circle, Srikakulam
APPSC Group-2 Geography Mock Test — Flask backend

Run:
    pip install -r requirements.txt
    python app.py
Then open http://127.0.0.1:5000
"""

import json
import os
import sqlite3
import uuid
import re
from datetime import datetime

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, send_file, abort
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "mocktest.db")
QUESTIONS_PATH = os.path.join(BASE_DIR, "questions.json")
RESULTS_XLSX = os.path.join(BASE_DIR, "results.xlsx")

app = Flask(__name__)
# Change this secret key before any real deployment.
app.secret_key = os.environ.get("SECRET_KEY", "jphule-srikakulam-group2-geo-2026-change-me")

# ----------------------------------------------------------------------------
# Question bank
# ----------------------------------------------------------------------------
with open(QUESTIONS_PATH, "r", encoding="utf-8") as f:
    _QDATA = json.load(f)

META = _QDATA["meta"]
QUESTIONS = _QDATA["questions"]
# Fast lookup of correct answers by id (never sent to the client)
ANSWER_KEY = {q["id"]: q["correct"] for q in QUESTIONS}
TOPIC_BY_ID = {q["id"]: q.get("topic", "General") for q in QUESTIONS}

DURATION_MINUTES = META.get("duration_minutes", 150)
NEG_FRACTION = 1.0 / 3.0  # 1/3 negative marking

# Client-safe copy of questions (correct answer + explanation stripped)
def _client_questions():
    out = []
    for q in QUESTIONS:
        out.append({
            "id": q["id"],
            "topic": q.get("topic", ""),
            "question": q["question"],
            "options": q["options"],
        })
    return out

CLIENT_QUESTIONS = _client_questions()

# ----------------------------------------------------------------------------
# Database
# ----------------------------------------------------------------------------
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS students (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT NOT NULL,
            mobile        TEXT NOT NULL UNIQUE,
            email         TEXT NOT NULL,
            aadhar        TEXT,
            token         TEXT UNIQUE,
            registered_at TEXT,
            submitted     INTEGER DEFAULT 0,
            submitted_at  TEXT,
            score         REAL,
            correct       INTEGER,
            wrong         INTEGER,
            unattempted   INTEGER,
            answers_json  TEXT
        )
        """
    )
    conn.commit()
    conn.close()


init_db()

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
MOBILE_RE = re.compile(r"^\d{10}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def student_by_mobile(mobile):
    conn = get_db()
    row = conn.execute("SELECT * FROM students WHERE mobile = ?", (mobile,)).fetchone()
    conn.close()
    return row


def student_by_token(token):
    conn = get_db()
    row = conn.execute("SELECT * FROM students WHERE token = ?", (token,)).fetchone()
    conn.close()
    return row


def export_to_xlsx(student_row):
    """Append/update this student's result in results.xlsx."""
    if Workbook is None:
        return
    headers = [
        "Name", "Mobile", "Email", "Aadhar/ID",
        "Registered At", "Submitted At",
        "Score", "Correct", "Wrong", "Unattempted", "Total Questions",
    ]
    if os.path.exists(RESULTS_XLSX):
        wb = load_workbook(RESULTS_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(headers)

    ws.append([
        student_row["name"],
        student_row["mobile"],
        student_row["email"],
        student_row["aadhar"] or "",
        student_row["registered_at"],
        student_row["submitted_at"],
        student_row["score"],
        student_row["correct"],
        student_row["wrong"],
        student_row["unattempted"],
        META.get("total_questions", len(QUESTIONS)),
    ])
    wb.save(RESULTS_XLSX)


# ----------------------------------------------------------------------------
# Routes
# ----------------------------------------------------------------------------
@app.route("/")
def index():
    return redirect(url_for("register"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        mobile = (request.form.get("mobile") or "").strip()
        email = (request.form.get("email") or "").strip()
        aadhar = (request.form.get("aadhar") or "").strip()

        errors = []
        if not name:
            errors.append("దయచేసి మీ పూర్తి పేరు నమోదు చేయండి.")
        if not MOBILE_RE.match(mobile):
            errors.append("మొబైల్ నంబర్ సరిగ్గా 10 అంకెలు ఉండాలి.")
        if not EMAIL_RE.match(email):
            errors.append("దయచేసి సరైన ఇమెయిల్ చిరునామా నమోదు చేయండి.")

        # Single-attempt enforcement
        existing = student_by_mobile(mobile) if MOBILE_RE.match(mobile) else None
        if existing is not None:
            errors.append("మీరు ఇప్పటికే ఈ పరీక్షను రాశారు. ఒక్కసారి మాత్రమే ప్రయత్నించవచ్చు.")

        if errors:
            return render_template(
                "register.html", meta=META, errors=errors,
                form={"name": name, "mobile": mobile, "email": email, "aadhar": aadhar},
            )

        token = uuid.uuid4().hex
        conn = get_db()
        try:
            conn.execute(
                "INSERT INTO students (name, mobile, email, aadhar, token, registered_at) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (name, mobile, email, aadhar, token, datetime.now().isoformat(timespec="seconds")),
            )
            conn.commit()
        except sqlite3.IntegrityError:
            conn.close()
            return render_template(
                "register.html", meta=META,
                errors=["మీరు ఇప్పటికే ఈ పరీక్షను రాశారు. ఒక్కసారి మాత్రమే ప్రయత్నించవచ్చు."],
                form={"name": name, "mobile": mobile, "email": email, "aadhar": aadhar},
            )
        conn.close()

        session.clear()
        session["token"] = token
        return redirect(url_for("instructions"))

    return render_template("register.html", meta=META, errors=None, form={})


@app.route("/instructions")
def instructions():
    token = session.get("token")
    student = student_by_token(token) if token else None
    if student is None:
        return redirect(url_for("register"))
    if student["submitted"]:
        return redirect(url_for("result"))
    return render_template("instructions.html", meta=META)


@app.route("/test")
def test():
    token = session.get("token")
    student = student_by_token(token) if token else None
    if student is None:
        return redirect(url_for("register"))
    if student["submitted"]:
        return redirect(url_for("result"))
    return render_template(
        "test.html",
        meta=META,
        student_name=student["name"],
        duration_minutes=DURATION_MINUTES,
    )


@app.route("/api/questions")
def api_questions():
    token = session.get("token")
    student = student_by_token(token) if token else None
    if student is None:
        return jsonify({"error": "unauthorized"}), 401
    if student["submitted"]:
        return jsonify({"error": "already_submitted"}), 403
    return jsonify({
        "questions": CLIENT_QUESTIONS,
        "duration_minutes": DURATION_MINUTES,
        "total_questions": META.get("total_questions", len(QUESTIONS)),
    })


@app.route("/api/submit", methods=["POST"])
def api_submit():
    token = session.get("token")
    student = student_by_token(token) if token else None
    if student is None:
        return jsonify({"error": "unauthorized"}), 401
    if student["submitted"]:
        return jsonify({"error": "already_submitted"}), 403

    payload = request.get_json(silent=True) or {}
    answers = payload.get("answers", {})  # {"1": "A", "2": "C", ...}

    correct = wrong = unattempted = 0
    per_topic = {}
    for q in QUESTIONS:
        qid = q["id"]
        topic = TOPIC_BY_ID.get(qid, "General")
        per_topic.setdefault(topic, {"correct": 0, "wrong": 0, "unattempted": 0})
        given = answers.get(str(qid))
        if given is None or given == "":
            unattempted += 1
            per_topic[topic]["unattempted"] += 1
        elif given == ANSWER_KEY[qid]:
            correct += 1
            per_topic[topic]["correct"] += 1
        else:
            wrong += 1
            per_topic[topic]["wrong"] += 1

    score = round(correct - (wrong * NEG_FRACTION), 2)

    conn = get_db()
    conn.execute(
        "UPDATE students SET submitted=1, submitted_at=?, score=?, correct=?, wrong=?, "
        "unattempted=?, answers_json=? WHERE token=?",
        (
            datetime.now().isoformat(timespec="seconds"),
            score, correct, wrong, unattempted,
            json.dumps(answers, ensure_ascii=False), token,
        ),
    )
    conn.commit()
    conn.close()

    # Export to Excel
    try:
        export_to_xlsx(student_by_token(token))
    except Exception as exc:  # pragma: no cover
        app.logger.warning("xlsx export failed: %s", exc)

    return jsonify({"ok": True, "redirect": url_for("result")})


@app.route("/result")
def result():
    token = session.get("token")
    student = student_by_token(token) if token else None
    if student is None:
        return redirect(url_for("register"))
    if not student["submitted"]:
        return redirect(url_for("instructions"))

    total = META.get("total_questions", len(QUESTIONS))
    return render_template(
        "result.html",
        meta=META,
        student=student,
        total=total,
        answered=len(QUESTIONS),
    )


@app.route("/download-results")
def download_results():
    """Simple protected download of the results workbook.
    Set ADMIN_KEY env var and open /download-results?key=YOURKEY
    """
    admin_key = os.environ.get("ADMIN_KEY", "admin123")
    if request.args.get("key") != admin_key:
        abort(403)
    if not os.path.exists(RESULTS_XLSX):
        abort(404)
    return send_file(RESULTS_XLSX, as_attachment=True, download_name="results.xlsx")


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
